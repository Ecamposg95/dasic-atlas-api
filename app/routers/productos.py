from decimal import Decimal, InvalidOperation
from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, File
from fastapi.responses import Response, StreamingResponse
from sqlalchemy import func, or_, text
from sqlalchemy.orm import Session
from typing import List, Optional, Union
import csv
import io
import json
import secrets
from datetime import datetime
from pathlib import Path
from app import models
from app import schemas
from app.db import get_db
from app.models.enums import TipoMovimientoStock
from app.security import allow_admin, allow_admin_asistente, allow_all_staff, get_current_user
from app.services.stock_service import aplicar_movimiento

router = APIRouter(prefix="/api/productos", tags=["Productos"])

DEFAULT_PURCHASE_CURRENCY = "MXN"
DEFAULT_UNIDAD = "PZA"
MARCA_TAXONOMY_FILE = Path(__file__).resolve().parent.parent / "data" / "marca_abreviaturas.json"
IMPORT_COLUMN_ALIASES = {
    "sku": ("sku", "codigo", "codigo_interno"),
    "sku_comercial": ("sku_comercial", "sku comercial", "commercial_sku", "catalogo", "catálogo"),
    "nombre": ("nombre", "producto"),
    "descripcion": ("descripcion", "descripción"),
    "marca": ("marca", "brand"),
    "unidad": ("unidad", "unit"),
    "stock_actual": ("stock", "stock_actual", "existencia", "cantidad"),
    "stock_minimo": ("stock_minimo", "stock mínimo", "minimo"),
    "costo_compra": ("costo", "costo_compra", "purchase_cost", "precio unitario"),
    "precio_publico": ("precio_publico", "precio público", "precio_lista"),
    "precio_mayorista": ("precio_mayorista",),
    "precio_distribuidor": ("precio_distribuidor",),
    "moneda_compra": ("moneda_compra", "moneda", "currency"),
}
IMPORT_TEMPLATE_COLUMNS = [
    "sku",
    "sku_comercial",
    "nombre",
    "descripcion",
    "marca",
    "unidad",
    "stock",
    "stock_minimo",
    "costo",
    "moneda_compra",
    "precio_publico",
    "precio_mayorista",
    "precio_distribuidor",
]


def _normalize_sku(value: str) -> str:
    return value.strip().upper()


def _resolve_marca(db: Session, *, marca_id: Optional[int], marca_texto: Optional[str]) -> tuple[Optional[int], Optional[str]]:
    """Normaliza la entrada de marca a (marca_id_final, marca_texto_final).

    Reglas:
    - Si marca_id viene, debe existir → si no, 400.
    - Si solo marca_texto viene y matchea case-insensitive una fila de `marcas`,
      autocompleta marca_id.
    - Si solo marca_texto viene y no matchea, marca_id=None y se preserva el
      texto (compat con CSV legacy).
    """
    if marca_id is not None:
        m = db.get(models.Marca, marca_id)
        if not m:
            raise HTTPException(status_code=400, detail=f"marca_id={marca_id} no existe")
        return m.id, m.nombre

    texto = _normalize_optional_text(marca_texto)
    if not texto:
        return None, None

    m = (
        db.query(models.Marca)
        .filter(func.lower(models.Marca.nombre) == texto.lower())
        .first()
    )
    if m:
        return m.id, m.nombre
    return None, texto


def _validar_fk_proveedores(db: Session, *, principal_id, alterno_id) -> None:
    """Levanta 400 si algún proveedor_id viene no-None y no existe en la tabla.

    Antes era posible crear/editar producto con proveedor_id apuntando a una
    fila inexistente; la FK lo permitía técnicamente (default NO ACTION) pero
    quedaba huérfano en la app.
    """
    for nombre, pid in (("proveedor_principal_id", principal_id), ("proveedor_alterno_id", alterno_id)):
        if pid is None:
            continue
        if not db.get(models.Proveedor, pid):
            raise HTTPException(status_code=400, detail=f"{nombre}={pid} no existe")


def _generate_internal_sku(db: Session, marca: str | None = None, marca_id: int | None = None) -> str:
    """Genera SKU interno único.

    Resuelve la marca primero por `marca_id` (FK directa), luego por `marca`
    texto (match case-insensitive contra `marcas.nombre`). Si encuentra,
    delega en `siguiente_sku_para` (consecutivo `{ABREV}-{NNNN}` con advisory
    lock por abreviatura — serializa creadores concurrentes).

    Fallback: `DAS-YYMM-XXXX` (random hex) cuando no hay marca conocida.
    """
    marca_row = None
    if marca_id is not None:
        marca_row = db.get(models.Marca, marca_id)
    elif marca:
        marca_row = (
            db.query(models.Marca)
            .filter(func.lower(models.Marca.nombre) == marca.strip().lower())
            .first()
        )

    if marca_row:
        abrev = marca_row.abreviatura
        # Advisory lock por abreviatura: serializa contra otros creadores.
        db.execute(
            text("SELECT pg_advisory_xact_lock(hashtext(:k))"),
            {"k": f"sku:{abrev}"},
        )
        # Import dentro de función para evitar ciclo módulo.
        from app.routers.catalogos import siguiente_sku_para
        return siguiente_sku_para(db, abrev)

    # Legacy fallback para productos sin marca registrada.
    for _ in range(10):
        candidate = f"DAS-{datetime.utcnow().strftime('%y%m')}-{secrets.token_hex(2).upper()}"
        if not db.query(models.Producto).filter(models.Producto.sku == candidate).first():
            return candidate
    raise HTTPException(500, "No se pudo generar SKU único, reintentar")


def _read_xlsx_rows(content: bytes) -> List[dict]:
    """Lee un .xlsx y devuelve una lista de dicts (header → valor) por fila."""
    try:
        from openpyxl import load_workbook  # type: ignore
    except ImportError as exc:  # pragma: no cover
        raise HTTPException(
            400,
            detail={
                "mensaje": "Falta dependencia openpyxl en el server. Instala 'openpyxl' o usa CSV.",
                "usa_csv_mientras_tanto": True,
                "columnas_compatibles": IMPORT_TEMPLATE_COLUMNS,
            },
        ) from exc

    wb = load_workbook(io.BytesIO(content), data_only=True, read_only=True)
    ws = wb.active
    rows = ws.iter_rows(values_only=True)
    try:
        header = next(rows)
    except StopIteration:
        return []
    headers = [str(h).strip().lower() if h is not None else "" for h in header]
    out: List[dict] = []
    for raw in rows:
        if raw is None:
            continue
        if all(v is None or str(v).strip() == "" for v in raw):
            continue
        row = {}
        for idx, value in enumerate(raw):
            if idx >= len(headers):
                break
            row[headers[idx]] = value
        out.append(row)
    return out


def _normalize_optional_text(value: Optional[str]) -> Optional[str]:
    if value is None:
        return None
    normalized = value.strip()
    return normalized or None


def _normalize_currency(value: Optional[str]) -> str:
    normalized = (value or DEFAULT_PURCHASE_CURRENCY).strip().upper()
    return normalized or DEFAULT_PURCHASE_CURRENCY


def _resolve_public_price(raw_price: Optional[Decimal], costo_compra: Decimal) -> Optional[Decimal]:
    return raw_price


def _read_decimal(row: dict, field_name: str) -> Optional[Decimal]:
    for alias in IMPORT_COLUMN_ALIASES[field_name]:
        raw_value = row.get(alias)
        if raw_value is None:
            continue
        cleaned = str(raw_value).strip().replace(",", "")
        if cleaned == "":
            return None
        try:
            return Decimal(cleaned)
        except InvalidOperation as exc:
            raise ValueError(f"{field_name} inválido: {raw_value}") from exc
    return None


def _read_int(row: dict, field_name: str, default: int = 0) -> int:
    for alias in IMPORT_COLUMN_ALIASES[field_name]:
        raw_value = row.get(alias)
        if raw_value is None:
            continue
        cleaned = str(raw_value).strip()
        if cleaned == "":
            return default
        try:
            return int(Decimal(cleaned))
        except InvalidOperation as exc:
            raise ValueError(f"{field_name} inválido: {raw_value}") from exc
    return default


def _read_text(row: dict, field_name: str, default: Optional[str] = None) -> Optional[str]:
    for alias in IMPORT_COLUMN_ALIASES[field_name]:
        raw_value = row.get(alias)
        if raw_value is None:
            continue
        normalized = _normalize_optional_text(str(raw_value))
        if normalized is not None:
            return normalized
    return default


def _normalize_csv_row(row: dict) -> dict:
    return {
        (str(key).strip().lower() if key is not None else ""): (value.strip() if isinstance(value, str) else value)
        for key, value in row.items()
        if key is not None
    }

# --- 1. OBTENER PRODUCTOS (LISTADO INTELIGENTE) ---
@router.get("/", response_model=Union[List[schemas.ProductoResponseAdmin], List[schemas.ProductoResponseVendedor]])
def listar_productos(
    skip: int = Query(0, ge=0),
    limit: int = Query(500, ge=1, le=2000),
    q: Optional[str] = None,
    marca: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user: models.Usuario = Depends(allow_all_staff),
):
    """Lista productos. Admin/Asistente ven costo; vendedor solo precios de venta."""
    query = db.query(models.Producto)
    if q:
        like = f"%{q.strip()}%"
        query = query.filter(or_(
            models.Producto.sku.ilike(like),
            models.Producto.sku_comercial.ilike(like),
            models.Producto.nombre.ilike(like),
            models.Producto.marca.ilike(like),
        ))
    if marca:
        query = query.filter(models.Producto.marca.ilike(marca.strip()))

    productos = (
        query.order_by(models.Producto.nombre.asc())
        .offset(skip)
        .limit(limit)
        .all()
    )

    if current_user.rol in [models.RolUsuario.ADMIN, models.RolUsuario.ASISTENTE]:
        return [schemas.ProductoResponseAdmin.model_validate(p) for p in productos]
    return [schemas.ProductoResponseVendedor.model_validate(p) for p in productos]


# --- 1.5 LISTA DE MARCAS (taxonomía DASIC + marcas en uso) ---
@router.get("/marcas", dependencies=[Depends(allow_all_staff)])
def listar_marcas(db: Session = Depends(get_db)):
    """Devuelve marcas en uso + taxonomía histórica DASIC.

    Cada item: { abreviatura, marca, categoria, prefijo_sku, codigos_ejemplo }.
    El frontend puede usar `prefijo_sku` (ej. ABPB) para sugerir SKUs internos.
    """
    items_taxonomia: list[dict] = []
    if MARCA_TAXONOMY_FILE.exists():
        try:
            data = json.loads(MARCA_TAXONOMY_FILE.read_text(encoding="utf-8"))
            for it in data.get("items", []):
                items_taxonomia.append({
                    "abreviatura": it.get("abreviatura"),
                    "marca": (it.get("marca") or "").strip() or None,
                    "categoria": (it.get("categoria") or "").strip() or None,
                    "prefijo_sku": it.get("abreviatura"),
                    "codigos_ejemplo": it.get("codigos_ejemplo") or [],
                    "fuente": "taxonomia",
                })
        except Exception:
            pass

    en_uso = (
        db.query(models.Producto.marca)
        .filter(models.Producto.marca.is_not(None))
        .distinct()
        .all()
    )
    marcas_db = {m.strip() for (m,) in en_uso if m and m.strip()}
    marcas_taxonomia = {
        (i["marca"] or "").strip().upper() for i in items_taxonomia if i["marca"]
    }
    extras = [
        {
            "abreviatura": None,
            "marca": m,
            "categoria": None,
            "prefijo_sku": None,
            "codigos_ejemplo": [],
            "fuente": "db",
        }
        for m in sorted(marcas_db)
        if m.upper() not in marcas_taxonomia
    ]

    return {
        "items": items_taxonomia + extras,
        "unidades_sugeridas": ["PZA", "CAJA", "MTS", "KG", "JUEGO", "SERVICIO", "PAR"],
        "monedas": ["MXN", "USD"],
    }

# --- 2. OBTENER UN SOLO PRODUCTO ---
@router.get("/{id}", response_model=Union[schemas.ProductoResponseAdmin, schemas.ProductoResponseVendedor])
def obtener_producto(
    id: int, 
    db: Session = Depends(get_db),
    current_user: models.Usuario = Depends(allow_all_staff)
):
    producto = db.query(models.Producto).filter(models.Producto.id == id).first()
    if not producto:
        raise HTTPException(status_code=404, detail="Producto no encontrado")

    if current_user.rol in [models.RolUsuario.ADMIN, models.RolUsuario.ASISTENTE]:
        return schemas.ProductoResponseAdmin.model_validate(producto)
    return schemas.ProductoResponseVendedor.model_validate(producto)

# --- 3. CREAR PRODUCTO (SOLO ADMIN/ASISTENTE) ---
@router.post("/", response_model=schemas.ProductoResponseAdmin, dependencies=[Depends(allow_admin_asistente)])
def crear_producto(
    producto: schemas.ProductoCreate,
    db: Session = Depends(get_db),
    current_user: models.Usuario = Depends(get_current_user),
):
    _validar_fk_proveedores(
        db,
        principal_id=producto.proveedor_principal_id,
        alterno_id=producto.proveedor_alterno_id,
    )

    # Normaliza marca (texto + id → marca_id final + texto autocompletado).
    marca_id_final, marca_texto_final = _resolve_marca(
        db,
        marca_id=producto.marca_id,
        marca_texto=producto.marca,
    )

    raw_sku = (producto.sku or "").strip()
    sku = (
        _normalize_sku(raw_sku)
        if raw_sku
        else _generate_internal_sku(db, marca=marca_texto_final, marca_id=marca_id_final)
    )

    # Verificar si el SKU ya existe
    db_producto = db.query(models.Producto).filter(models.Producto.sku == sku).first()
    if db_producto:
        raise HTTPException(status_code=400, detail=f"El SKU '{sku}' ya existe.")

    payload = producto.model_dump()
    payload["sku"] = sku
    sku_comercial_norm = _normalize_optional_text(producto.sku_comercial)
    payload["sku_comercial"] = sku_comercial_norm or sku  # autocompleta con interno
    payload["moneda_compra"] = _normalize_currency(producto.moneda_compra)
    payload["marca"] = marca_texto_final
    payload["marca_id"] = marca_id_final
    payload["unidad"] = _normalize_optional_text(producto.unidad) or DEFAULT_UNIDAD
    if payload.get("precio_publico") is None:
        payload.pop("precio_publico", None)

    # Normaliza claves SAT (uppercase) y categoría
    if payload.get("clave_unidad_sat"):
        payload["clave_unidad_sat"] = payload["clave_unidad_sat"].strip().upper()
    if payload.get("clave_prod_serv"):
        payload["clave_prod_serv"] = payload["clave_prod_serv"].strip()
    if payload.get("categoria"):
        payload["categoria"] = payload["categoria"].strip()

    # El stock inicial entra al producto vía aplicar_movimiento (ENTRADA auditable).
    # Nace con 0 y se ajusta abajo, garantizando kardex completo desde día 1.
    stock_inicial = int(payload.pop("stock_actual", 0) or 0)
    payload["stock_actual"] = 0

    nuevo_producto = models.Producto(**payload)
    db.add(nuevo_producto)
    db.flush()

    if stock_inicial > 0:
        aplicar_movimiento(
            db,
            producto=nuevo_producto,
            tipo=TipoMovimientoStock.ENTRADA.value,
            cantidad=stock_inicial,
            referencia_tipo="stock_inicial",
            motivo="alta de producto con stock inicial",
            usuario=current_user,
        )

    db.commit()
    db.refresh(nuevo_producto)
    return nuevo_producto

# --- 4. ACTUALIZAR PRODUCTO ---
@router.put("/{id}", response_model=schemas.ProductoResponseAdmin, dependencies=[Depends(allow_admin_asistente)])
def actualizar_producto(
    id: int,
    producto_update: schemas.ProductoUpdate,
    db: Session = Depends(get_db),
    current_user: models.Usuario = Depends(get_current_user),
):
    db_producto = db.query(models.Producto).filter(models.Producto.id == id).first()
    if not db_producto:
        raise HTTPException(status_code=404, detail="Producto no encontrado")

    _validar_fk_proveedores(
        db,
        principal_id=producto_update.proveedor_principal_id,
        alterno_id=producto_update.proveedor_alterno_id,
    )

    # Actualizamos solo los campos enviados.
    update_data = producto_update.model_dump(exclude_unset=True)
    # stock_actual sale del setattr loop: lo procesamos al final como AJUSTE auditable.
    nuevo_stock = update_data.pop("stock_actual", None)

    if "sku_comercial" in update_data:
        update_data["sku_comercial"] = _normalize_optional_text(update_data["sku_comercial"])

    # Si el body trajo marca y/o marca_id, normalizamos ambos juntos.
    if "marca" in update_data or "marca_id" in update_data:
        marca_id_in = update_data.get("marca_id", db_producto.marca_id)
        marca_texto_in = update_data.get("marca", db_producto.marca)
        marca_id_final, marca_texto_final = _resolve_marca(
            db, marca_id=marca_id_in, marca_texto=marca_texto_in
        )
        update_data["marca_id"] = marca_id_final
        update_data["marca"] = marca_texto_final

    if "unidad" in update_data:
        update_data["unidad"] = _normalize_optional_text(update_data["unidad"]) or DEFAULT_UNIDAD
    if "moneda_compra" in update_data:
        update_data["moneda_compra"] = _normalize_currency(update_data["moneda_compra"])

    # Normalizaciones SAT + clasificación
    if "clave_unidad_sat" in update_data and update_data["clave_unidad_sat"]:
        update_data["clave_unidad_sat"] = update_data["clave_unidad_sat"].strip().upper()
    if "clave_prod_serv" in update_data and update_data["clave_prod_serv"]:
        update_data["clave_prod_serv"] = update_data["clave_prod_serv"].strip()
    if "categoria" in update_data and update_data["categoria"]:
        update_data["categoria"] = update_data["categoria"].strip()

    for key, value in update_data.items():
        setattr(db_producto, key, value)

    # Si el body trajo stock_actual y cambia, emitimos un AJUSTE por el delta.
    # Esto preserva kardex aún cuando el cambio entra por el PUT y no por /ajustar-stock.
    if nuevo_stock is not None:
        delta = int(nuevo_stock) - int(db_producto.stock_actual or 0)
        if delta != 0:
            try:
                aplicar_movimiento(
                    db,
                    producto=db_producto,
                    tipo=TipoMovimientoStock.AJUSTE.value,
                    cantidad=delta,
                    referencia_tipo="put_producto",
                    motivo="ajuste vía edición de producto",
                    usuario=current_user,
                )
            except ValueError as exc:
                db.rollback()
                raise HTTPException(status_code=400, detail=str(exc))

    db.commit()
    db.refresh(db_producto)
    return db_producto

# --- 5. ELIMINAR PRODUCTO (SOLO ADMIN) ---
@router.delete("/{id}", dependencies=[Depends(allow_admin)])
def eliminar_producto(id: int, db: Session = Depends(get_db)):
    db_producto = db.query(models.Producto).filter(models.Producto.id == id).first()
    if not db_producto:
        raise HTTPException(status_code=404, detail="Producto no encontrado")

    # Kardex es inmutable. La DB tiene ON DELETE RESTRICT, pero damos un
    # error explicativo antes de que Postgres regrese SQLSTATE 23503.
    kardex = (
        db.query(models.MovimientoStock.id)
        .filter(models.MovimientoStock.producto_id == id)
        .first()
    )
    if kardex:
        raise HTTPException(
            status_code=409,
            detail=(
                "El producto tiene historial en el kardex. No se puede eliminar. "
                "Considera marcarlo como inactivo (TODO) o ajustar su stock a 0."
            ),
        )

    # detalles_orden / detalles_compra ahora son ON DELETE SET NULL: el
    # producto puede borrarse y los detalles históricos conservan sku_libre.
    db.delete(db_producto)
    db.commit()
    return {"mensaje": "Producto eliminado correctamente"}


# --- 5.5 AJUSTAR STOCK MANUALMENTE (Admin/Asistente) ---
from pydantic import BaseModel as _PydBase  # noqa: E402


class AjusteStockIn(_PydBase):
    delta: int
    motivo: Optional[str] = None


@router.post("/{id}/ajustar-stock", response_model=schemas.ProductoResponseAdmin, dependencies=[Depends(allow_admin_asistente)])
def ajustar_stock(
    id: int,
    payload: AjusteStockIn,
    db: Session = Depends(get_db),
    current_user: models.Usuario = Depends(get_current_user),
):
    """Suma `delta` al stock_actual (puede ser negativo). El stock no baja de 0.

    Pasa por aplicar_movimiento para que la mutación quede auditada en
    MovimientoStock (kardex). Antes mutaba stock_actual directamente.
    """
    producto = db.get(models.Producto, id)
    if not producto:
        raise HTTPException(status_code=404, detail="Producto no encontrado")
    try:
        aplicar_movimiento(
            db,
            producto=producto,
            tipo=TipoMovimientoStock.AJUSTE.value,
            cantidad=payload.delta,
            referencia_tipo="ajuste_manual",
            motivo=payload.motivo or "ajuste manual vía /productos/ajustar-stock",
            usuario=current_user,
        )
        db.commit()
        db.refresh(producto)
        return producto
    except ValueError as exc:
        db.rollback()
        raise HTTPException(status_code=400, detail=str(exc))

# --- 6. CARGA MASIVA CSV / XLSX (IMPORTAR) ---
@router.post("/upload-csv", dependencies=[Depends(allow_admin_asistente)])
async def cargar_inventario_csv(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: models.Usuario = Depends(get_current_user),
):
    """
    Importa productos desde CSV o Excel (.xlsx).
    Columnas base: sku (opcional, se autogenera), nombre, costo, stock.
    Columnas opcionales: sku_comercial, moneda_compra, precio_publico,
    precio_mayorista, precio_distribuidor, descripcion, stock_minimo.

    Semántica de stock: REPLACE auditable. Si el SKU ya existe, el stock del
    CSV reemplaza al actual (no se suma). El delta queda registrado en kardex
    como MovimientoStock AJUSTE con referencia_tipo='import_csv'. Re-subir el
    mismo archivo no duplica stock.
    """
    filename = (file.filename or "").lower()

    # Límite de tamaño 25 MB. CSV/XLSX de inventario reales pesan KB-MB; 25 MB
    # cubre catálogos enormes (~100k filas). Más allá = DoS por memoria.
    MAX_UPLOAD_BYTES = 25 * 1024 * 1024
    if file.size is not None and file.size > MAX_UPLOAD_BYTES:
        raise HTTPException(
            status_code=413,
            detail=f"Archivo demasiado grande ({file.size} bytes). Máximo {MAX_UPLOAD_BYTES} bytes (25 MB).",
        )
    content = await file.read()
    if len(content) > MAX_UPLOAD_BYTES:
        raise HTTPException(
            status_code=413,
            detail=f"Archivo demasiado grande ({len(content)} bytes). Máximo {MAX_UPLOAD_BYTES} bytes (25 MB).",
        )

    if filename.endswith(".xls"):
        raise HTTPException(
            status_code=400,
            detail={
                "mensaje": "Formato .xls no soportado. Convierte a .xlsx o CSV.",
                "columnas_compatibles": IMPORT_TEMPLATE_COLUMNS,
            },
        )

    if filename.endswith(".xlsx"):
        rows_iter = _read_xlsx_rows(content)
    elif filename.endswith(".csv") or not filename:
        decoded_content = content.decode('utf-8-sig')
        rows_iter = list(csv.DictReader(io.StringIO(decoded_content)))
    else:
        raise HTTPException(status_code=400, detail="El archivo debe ser CSV o XLSX")

    registros_nuevos = 0
    registros_actualizados = 0
    registros_omitidos = 0
    errores = []

    for idx, row in enumerate(rows_iter, start=2):  # +1 por header, +1 base-1 humano
        try:
            row = _normalize_csv_row(row)

            sku = _read_text(row, "sku")
            nombre_check = _read_text(row, "nombre")
            if not sku and not nombre_check:
                continue  # Fila vacía

            # Leer campos ANTES de autogenerar SKU: marca participa en el
            # SKU autogenerado y antes estaba siendo referenciada sin definir.
            sku_comercial = _read_text(row, "sku_comercial")
            nombre = _read_text(row, "nombre", "Sin Nombre") or "Sin Nombre"
            descripcion = _read_text(row, "descripcion", "") or ""
            marca = _read_text(row, "marca")
            unidad = _read_text(row, "unidad") or DEFAULT_UNIDAD
            moneda_compra = _normalize_currency(_read_text(row, "moneda_compra", DEFAULT_PURCHASE_CURRENCY))
            costo_compra = _read_decimal(row, "costo_compra") or Decimal("0")
            precio_publico = _read_decimal(row, "precio_publico")
            precio_mayorista = _read_decimal(row, "precio_mayorista") or Decimal("0")
            precio_distribuidor = _read_decimal(row, "precio_distribuidor") or Decimal("0")
            stock_actual = _read_int(row, "stock_actual", 0)
            stock_minimo = _read_int(row, "stock_minimo", 5)

            # Resolver marca texto → (marca_id, texto normalizado) para
            # autocompletar la FK cuando el texto del CSV matchea una marca
            # registrada. Si no matchea, queda solo el texto (compat legacy).
            marca_id_csv, marca_texto_csv = _resolve_marca(
                db, marca_id=None, marca_texto=marca,
            )

            if not sku:
                # Autogenerar SKU interno si no viene en el archivo.
                sku = _generate_internal_sku(db, marca=marca_texto_csv, marca_id=marca_id_csv)
            else:
                sku = _normalize_sku(sku)

            if stock_actual < 0:
                raise ValueError(
                    f"stock_actual no puede ser negativo (recibido: {stock_actual})"
                )
            if stock_minimo < 0:
                raise ValueError(
                    f"stock_minimo no puede ser negativo (recibido: {stock_minimo})"
                )
            if costo_compra < 0:
                raise ValueError(
                    f"costo_compra no puede ser negativo (recibido: {costo_compra})"
                )

            # Buscar si existe para actualizar (Upsert)
            producto_existente = db.query(models.Producto).filter(models.Producto.sku == sku).first()

            if producto_existente:
                # Metadata directo (no stock).
                producto_existente.sku_comercial = sku_comercial or producto_existente.sku_comercial or sku
                producto_existente.nombre = nombre or producto_existente.nombre
                producto_existente.descripcion = descripcion or producto_existente.descripcion
                if marca_texto_csv:
                    producto_existente.marca = marca_texto_csv
                if marca_id_csv is not None or marca_texto_csv is None:
                    # Solo sobreescribimos marca_id si el CSV trae marca explícita
                    # (id resolved o vacío). No tocamos si el CSV no traía marca.
                    producto_existente.marca_id = marca_id_csv
                if unidad:
                    producto_existente.unidad = unidad
                producto_existente.stock_minimo = stock_minimo
                producto_existente.moneda_compra = moneda_compra
                producto_existente.costo_compra = costo_compra
                if precio_publico is not None:
                    producto_existente.precio_publico = precio_publico
                producto_existente.precio_mayorista = precio_mayorista
                producto_existente.precio_distribuidor = precio_distribuidor

                # Stock con semántica REPLACE auditable: el valor del CSV se
                # vuelve absoluto. Si difiere del actual, emitimos un AJUSTE
                # por el delta para preservar kardex. Re-subir el mismo CSV no
                # duplica nada (delta = 0 → sin movimiento).
                delta = stock_actual - int(producto_existente.stock_actual or 0)
                if delta != 0:
                    aplicar_movimiento(
                        db,
                        producto=producto_existente,
                        tipo=TipoMovimientoStock.AJUSTE.value,
                        cantidad=delta,
                        referencia_tipo="import_csv",
                        motivo=f"import {file.filename} fila {idx}",
                        usuario=current_user,
                    )

                registros_actualizados += 1
            else:
                # Crear nuevo con stock=0 y ENTRADA por stock inicial (kardex completo).
                kwargs = dict(
                    sku=sku,
                    sku_comercial=sku_comercial or sku,
                    nombre=nombre,
                    descripcion=descripcion,
                    marca=marca_texto_csv,
                    marca_id=marca_id_csv,
                    unidad=unidad,
                    precio_mayorista=precio_mayorista,
                    precio_distribuidor=precio_distribuidor,
                    costo_compra=costo_compra,
                    moneda_compra=moneda_compra,
                    stock_actual=0,
                    stock_minimo=stock_minimo,
                )
                if precio_publico is not None:
                    kwargs["precio_publico"] = precio_publico
                nuevo_prod = models.Producto(**kwargs)
                db.add(nuevo_prod)
                db.flush()
                if stock_actual > 0:
                    aplicar_movimiento(
                        db,
                        producto=nuevo_prod,
                        tipo=TipoMovimientoStock.ENTRADA.value,
                        cantidad=stock_actual,
                        referencia_tipo="import_csv",
                        motivo=f"alta vía import {file.filename} fila {idx}",
                        usuario=current_user,
                    )
                registros_nuevos += 1

        except Exception as e:
            registros_omitidos += 1
            sku_ref = (row.get('sku') or row.get('SKU') or '?') if isinstance(row, dict) else '?'
            errores.append(f"Fila {idx} (SKU {sku_ref}): {str(e)}")

    db.commit()
    return {
        "mensaje": "Proceso completado",
        "creados": registros_nuevos,
        "actualizados": registros_actualizados,
        "omitidos": registros_omitidos,
        "errores": errores,
    }

# --- 6.5 QR DEL PRODUCTO (PNG) ---
@router.get("/{id}/qr", dependencies=[Depends(allow_all_staff)])
def producto_qr(id: int, db: Session = Depends(get_db)):
    producto = db.query(models.Producto).filter(models.Producto.id == id).first()
    if not producto:
        raise HTTPException(404, "Producto no encontrado")
    try:
        import qrcode  # type: ignore
    except ImportError:
        raise HTTPException(
            500,
            "Falta dependencia 'qrcode' en el server. Instala 'qrcode[pil]'.",
        )
    payload = (producto.sku_comercial or producto.sku or str(producto.id)).strip()
    img = qrcode.make(payload)
    buf = io.BytesIO()
    img.save(buf, format="PNG")
    return Response(
        content=buf.getvalue(),
        media_type="image/png",
        headers={"Content-Disposition": f"inline; filename=qr_{payload}.png"},
    )


# --- 7. EXPORTAR CSV (DESCARGAR) ---
@router.get("/exportar/csv", dependencies=[Depends(allow_admin_asistente)])
def exportar_productos_csv(db: Session = Depends(get_db)):
    """Genera y descarga un archivo CSV con todo el inventario"""
    productos = db.query(models.Producto).all()
    
    # Crear archivo en memoria
    output = io.StringIO()
    writer = csv.writer(output)
    
    # Encabezados
    writer.writerow([
        'SKU', 'SKU Comercial', 'Nombre', 'Descripción', 'Marca', 'Unidad',
        'Stock Actual', 'Stock Mínimo',
        'Costo', 'Moneda Compra', 'Precio Público', 'Precio Mayorista', 'Precio Distribuidor'
    ])

    # Datos
    for p in productos:
        writer.writerow([
            p.sku, p.sku_comercial, p.nombre, p.descripcion, p.marca, p.unidad,
            p.stock_actual, p.stock_minimo,
            p.costo_compra, p.moneda_compra, p.precio_publico, p.precio_mayorista, p.precio_distribuidor
        ])
    
    output.seek(0)

    # StreamingResponse permite descargar el archivo al vuelo
    response = StreamingResponse(iter([output.getvalue()]), media_type="text/csv")
    response.headers["Content-Disposition"] = "attachment; filename=inventario_dasic.csv"
    return response


# --- 8. CATEGORÍAS (distinct para autocomplete del modal) ---
@router.get("/utils/categorias", dependencies=[Depends(allow_all_staff)])
def listar_categorias(db: Session = Depends(get_db)):
    """Devuelve categorías distintas en uso + conteo. Para datalist del modal."""
    rows = (
        db.query(models.Producto.categoria, func.count(models.Producto.id).label("n"))
        .filter(models.Producto.categoria.is_not(None))
        .filter(models.Producto.categoria != "")
        .group_by(models.Producto.categoria)
        .order_by(models.Producto.categoria)
        .all()
    )
    return {"items": [{"categoria": c, "n_productos": int(n)} for (c, n) in rows]}


# --- 9. CARDEX (detalle completo: identificación + fiscales + movimientos + métricas) ---
@router.get("/{id}/cardex", dependencies=[Depends(allow_all_staff)])
def cardex_producto(id: int, db: Session = Depends(get_db)):
    """Vista expandida del producto: datos generales, datos fiscales SAT,
    últimos movimientos y métricas de uso (fechas, conteos). Usado por el
    side-panel de /inventario."""
    p = db.query(models.Producto).filter(models.Producto.id == id).first()
    if not p:
        raise HTTPException(404, "Producto no encontrado")

    # Últimos 90 días de movimientos (capeado a 100 filas para no inflar payload)
    movs = (
        db.query(models.MovimientoStock)
        .filter(models.MovimientoStock.producto_id == id)
        .order_by(models.MovimientoStock.creado_en.desc())
        .limit(100)
        .all()
    )

    # Métricas históricas
    primer_mov = (
        db.query(func.min(models.MovimientoStock.creado_en))
        .filter(models.MovimientoStock.producto_id == id).scalar()
    )
    ultimo_mov = (
        db.query(func.max(models.MovimientoStock.creado_en))
        .filter(models.MovimientoStock.producto_id == id).scalar()
    )
    n_movs = (
        db.query(func.count(models.MovimientoStock.id))
        .filter(models.MovimientoStock.producto_id == id).scalar() or 0
    )
    ultimo_uso_cot = (
        db.query(func.max(models.OrdenVenta.fecha_creacion))
        .join(models.DetalleOrden, models.DetalleOrden.orden_id == models.OrdenVenta.id)
        .filter(models.DetalleOrden.producto_id == id).scalar()
    )

    return {
        "identificacion": {
            "id": p.id,
            "sku": p.sku,                       # interno (= "abreviatura": ABCS-0001)
            "sku_comercial": p.sku_comercial,   # catálogo del fabricante
            "nombre": p.nombre,
            "descripcion": p.descripcion,
            "imagen_url": p.imagen_url,
            "es_servicio": bool(p.es_servicio),
        },
        "clasificacion": {
            "marca": p.marca,
            "marca_id": p.marca_id,
            "categoria": p.categoria,
            "unidad": p.unidad,
        },
        "fiscales": {
            "clave_prod_serv": p.clave_prod_serv,
            "clave_unidad_sat": p.clave_unidad_sat,
            "objeto_imp": p.objeto_imp,
            "descripcion_fiscal": p.descripcion_fiscal,
        },
        "inventario": {
            "stock_actual": int(p.stock_actual or 0),
            "stock_minimo": int(p.stock_minimo or 0),
            "moneda_compra": p.moneda_compra,
            "costo_compra": float(p.costo_compra or 0),
            "proveedor_principal_id": p.proveedor_principal_id,
            "proveedor_alterno_id": p.proveedor_alterno_id,
        },
        "historico": {
            "primer_movimiento": primer_mov.isoformat() if primer_mov else None,
            "ultimo_movimiento": ultimo_mov.isoformat() if ultimo_mov else None,
            "total_movimientos": int(n_movs),
            "ultimo_uso_en_cotizacion": ultimo_uso_cot.isoformat() if ultimo_uso_cot else None,
        },
        "movimientos": [
            {
                "id": m.id,
                "tipo": m.tipo,
                "cantidad": m.cantidad,
                "stock_resultante": m.stock_resultante,
                "referencia_tipo": m.referencia_tipo,
                "referencia_id": m.referencia_id,
                "motivo": m.motivo,
                "creado_en": m.creado_en.isoformat() if m.creado_en else None,
            }
            for m in movs
        ],
    }

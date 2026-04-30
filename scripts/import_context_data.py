"""
Carga inicial de datos a partir de los documentos de referencia en context/.

Fuentes:
  - context/LISTA_MATERIAL.xlsx              -> productos (inventario real)
  - context/01 AB_05 JUNIO_ 2022. REF.VTA (1).xlsx
                                             -> taxonomía marca/categoría -> app/data/marca_abreviaturas.json
  - context/C-2604227.pdf, C-2604229.pdf     -> clientes (Vitracoat)
  - context/OC-2604001.pdf                   -> proveedor (Dimeint)
  - context/R-2604001.pdf                    -> cliente (Vitracoat / Popoca)

Uso:
    python scripts/import_context_data.py [--dry-run]

Idempotente: si encuentra el mismo SKU / email lo deja como está.
"""

from __future__ import annotations

import argparse
import json
import logging
import sys
from datetime import datetime
from decimal import Decimal
from pathlib import Path
from typing import Iterable

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))

from openpyxl import load_workbook  # noqa: E402

from app.db import SessionLocal  # noqa: E402
from app import models  # noqa: E402

logging.basicConfig(level=logging.INFO, format="%(levelname)s %(message)s")
log = logging.getLogger("import_context")

CONTEXT_DIR = ROOT / "context"
DATA_DIR = ROOT / "app" / "data"

LISTA_MATERIAL_FILE = CONTEXT_DIR / "LISTA_MATERIAL.xlsx"
REF_VTA_FILE = CONTEXT_DIR / "01 AB_05 JUNIO_ 2022. REF.VTA (1).xlsx"

CURRENCY_MAP = {
    "MN": "MXN",
    "MXN": "MXN",
    "USD": "USD",
    "DLS": "USD",
}


# ---------- Productos ----------

def parse_lista_material(path: Path) -> list[dict]:
    wb = load_workbook(path, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    out = []
    for raw in rows[1:]:
        if not raw or all(c is None for c in raw):
            continue
        catalog, descripcion, marca, cantidad, unidad, precio, moneda = (
            list(raw) + [None] * 7
        )[:7]
        if not catalog:
            continue
        sku = str(catalog).strip()
        out.append({
            "sku": sku,
            "nombre": str(descripcion or sku).strip(),
            "marca": (str(marca).strip() if marca else None),
            "unidad": (str(unidad).strip() if unidad else "PZA"),
            "stock_actual": int(cantidad or 0),
            "costo_compra": Decimal(str(precio or 0)),
            "moneda_compra": CURRENCY_MAP.get(
                str(moneda or "MN").strip().upper(), "MXN"
            ),
        })
    return out


def upsert_productos(db, items: Iterable[dict], dry_run: bool) -> tuple[int, int]:
    created = updated = 0
    for it in items:
        existing = db.query(models.Producto).filter(models.Producto.sku == it["sku"]).first()
        if existing:
            existing.nombre = it["nombre"][:150]
            existing.descripcion = it["nombre"]
            existing.marca = (it["marca"] or None) and it["marca"][:80]
            existing.unidad = (it["unidad"] or "PZA")[:20]
            existing.costo_compra = it["costo_compra"]
            existing.moneda_compra = it["moneda_compra"]
            if not existing.sku_comercial:
                existing.sku_comercial = it["sku"][:80]
            if (existing.stock_actual or 0) == 0:
                existing.stock_actual = it["stock_actual"]
            updated += 1
        else:
            p = models.Producto(
                sku=it["sku"][:50],
                sku_comercial=it["sku"][:80],
                nombre=it["nombre"][:150],
                descripcion=it["nombre"],
                marca=(it["marca"] or None) and it["marca"][:80],
                unidad=(it["unidad"] or "PZA")[:20],
                costo_compra=it["costo_compra"],
                moneda_compra=it["moneda_compra"],
                stock_actual=it["stock_actual"],
                stock_minimo=5,
                precio_publico=Decimal("0.00"),
            )
            db.add(p)
            created += 1
    if not dry_run:
        db.commit()
    else:
        db.rollback()
    return created, updated


# ---------- Marcas / abreviaturas ----------

def parse_ref_vta(path: Path) -> list[dict]:
    wb = load_workbook(path, data_only=True)
    ws = wb["Sheet1"] if "Sheet1" in wb.sheetnames else wb.active
    out = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row:
            continue
        abrev, marca, categoria = row[0], row[1], row[2]
        if not abrev:
            continue
        codigos = [c for c in row[3:] if c]
        out.append({
            "abreviatura": str(abrev).strip(),
            "marca": str(marca).strip() if marca else None,
            "categoria": str(categoria).strip() if categoria else None,
            "codigos_ejemplo": [str(c).strip() for c in codigos],
        })
    return out


def write_marca_taxonomy(items: list[dict], dry_run: bool) -> Path:
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    out_path = DATA_DIR / "marca_abreviaturas.json"
    payload = {
        "fuente": "context/01 AB_05 JUNIO_ 2022. REF.VTA (1).xlsx",
        "descripcion": "Tabla histórica DASIC: prefijos para SKUs internos por marca/categoría.",
        "items": items,
    }
    if not dry_run:
        out_path.write_text(
            json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8"
        )
    return out_path


# ---------- Clientes / proveedores ----------

CLIENTES_SEED = [
    {
        "nombre_empresa": "Vitracoat Pinturas en Polvo",
        "contacto_nombre": "Luis Robles",
        "email": "auxcompras3@vitracoat.com",
        "telefono": None,
        "direccion": None,
    },
    {
        "nombre_empresa": "Vitracoat Pinturas en Polvo",
        "contacto_nombre": "Ing. José Popoca",
        "email": "jpopoca@vitracoat.com",
        "telefono": None,
        "direccion": None,
    },
]

PROVEEDORES_SEED = [
    {
        "nombre_empresa": "Dimeint",
        "contacto_nombre": "Arturo Ramírez",
        "email": "aramirez@dimeint.com",
        "telefono": None,
    },
]


def upsert_clientes(db, dry_run: bool) -> tuple[int, int]:
    created = skipped = 0
    for c in CLIENTES_SEED:
        existing = (
            db.query(models.Cliente)
            .filter(models.Cliente.email == c["email"])
            .first()
        )
        if existing:
            skipped += 1
            continue
        db.add(models.Cliente(**c))
        created += 1
    if not dry_run:
        db.commit()
    else:
        db.rollback()
    return created, skipped


def upsert_proveedores(db, dry_run: bool) -> tuple[int, int]:
    created = skipped = 0
    for p in PROVEEDORES_SEED:
        existing = (
            db.query(models.Proveedor)
            .filter(models.Proveedor.email == p["email"])
            .first()
        )
        if existing:
            skipped += 1
            continue
        db.add(models.Proveedor(**p))
        created += 1
    if not dry_run:
        db.commit()
    else:
        db.rollback()
    return created, skipped


# ---------- Documentos de muestra (cotizaciones / OC reales) ----------

def _admin_user(db) -> "models.Usuario | None":
    return (
        db.query(models.Usuario)
        .order_by(models.Usuario.id.asc())
        .first()
    )


def _cliente_por_email(db, email: str) -> "models.Cliente | None":
    return db.query(models.Cliente).filter(models.Cliente.email == email).first()


def _proveedor_por_email(db, email: str) -> "models.Proveedor | None":
    return db.query(models.Proveedor).filter(models.Proveedor.email == email).first()


def _ensure_producto(db, sku: str, nombre: str, costo: Decimal, moneda: str = "MXN"):
    p = db.query(models.Producto).filter(models.Producto.sku == sku).first()
    if p:
        if not p.sku_comercial:
            p.sku_comercial = sku[:80]
        return p
    p = models.Producto(
        sku=sku[:50],
        sku_comercial=sku[:80],
        nombre=nombre[:150],
        descripcion=nombre,
        costo_compra=costo,
        moneda_compra=moneda,
        stock_actual=0,
        stock_minimo=0,
        precio_publico=Decimal("0.00"),
    )
    db.add(p)
    db.flush()
    return p


# Datos de los PDFs reales en context/
COTIZACIONES_SEED = [
    {
        "folio": "C-2604227",
        "fecha": datetime(2026, 4, 22),
        "cliente_email": "auxcompras3@vitracoat.com",
        "moneda": "MXN",
        "lineas": [
            ("XB4BA31",  "BOTON PULSADOR RASANTE VERDE / T. E. S. P. V. 7-8 Días Habiles", 5, "283.38"),
            ("XB4BA42",  "BOTON PULSADOR RASANTE ROJO / T. E. S. P. V. 7-8 Días Habiles",  5, "283.38"),
            ("XB4BVG3",  "PILOTO LUMINOSO LED VERDE / T. E. S. P. V. 7-8 Días Habiles",    8, "276.80"),
            ("18435",    "LINTERNA PARA MINERO 18435 / T. E. S. P. V. 7-8 Días Habiles",   6, "762.75"),
        ],
    },
    {
        "folio": "C-2604229",
        "fecha": datetime(2026, 4, 24),
        "cliente_email": "jpopoca@vitracoat.com",
        "moneda": "MXN",
        "observaciones": "50% DE ANTICIPO Y 50% DE CRÉDITO.",
        "lineas": [
            (
                "SERVICIO",
                "Programación PLC y HMI para secuencia de arranque de Molino 15 en modo "
                "ciclón y modo Collector. Incluye programación de PLC, HMI y Variadores, "
                "más material listado en alcance (arrancador estado sólido, gabinetes, "
                "cables, clemas, contactores, interruptores, condulets, electroválvulas, "
                "etc.). / T. E. S. P. V.",
                1,
                "263000.00",
            ),
        ],
    },
]

ORDENES_COMPRA_SEED = [
    {
        "folio": "OC-2604001",
        "fecha": datetime(2026, 4, 16),
        "proveedor_email": "aramirez@dimeint.com",
        "moneda": "MXN",
        "lineas": [
            ("6ES7954-8LC04-0AA0", "SIMATIC S7 memory card, 4MB", 1, "1161.00"),
        ],
        "iva_porcentaje": Decimal("0.16"),
    },
]


def upsert_cotizaciones(db, dry_run: bool) -> tuple[int, int]:
    created = skipped = 0
    admin = _admin_user(db)
    if not admin:
        log.warning("No hay usuario admin; cotizaciones de muestra omitidas.")
        return 0, 0
    for c in COTIZACIONES_SEED:
        if db.query(models.OrdenVenta).filter(models.OrdenVenta.folio == c["folio"]).first():
            skipped += 1
            continue
        cliente = _cliente_por_email(db, c["cliente_email"])
        if not cliente:
            log.warning("Cliente %s no encontrado, salto %s", c["cliente_email"], c["folio"])
            continue

        total = Decimal("0.00")
        detalles = []
        for sku_libre, descripcion, qty, precio in c["lineas"]:
            precio_d = Decimal(precio)
            subtotal = (precio_d * qty).quantize(Decimal("0.01"))
            total += subtotal
            detalles.append(models.DetalleOrden(
                producto_id=None,
                sku_libre=sku_libre[:80],
                descripcion_libre=descripcion[:255],
                moneda_origen_linea=c["moneda"],
                costo_base_linea=precio_d,
                cantidad=qty,
                precio_unitario=precio_d,
                utilidad_aplicada=Decimal("0.00"),
                descuento_aplicado=Decimal("0.00"),
                subtotal=subtotal,
            ))

        orden = models.OrdenVenta(
            folio=c["folio"],
            cliente_id=cliente.id,
            vendedor_id=admin.id,
            fecha_creacion=c["fecha"],
            estatus=models.EstatusOrden.COTIZACION,
            moneda=c["moneda"],
            tipo_cambio=Decimal("1.0"),
            total=total,
            observaciones=c.get("observaciones"),
            version=1,
            detalles=detalles,
        )
        db.add(orden)
        created += 1
    if not dry_run:
        db.commit()
    else:
        db.rollback()
    return created, skipped


def upsert_ordenes_compra(db, dry_run: bool) -> tuple[int, int]:
    created = skipped = 0
    for oc in ORDENES_COMPRA_SEED:
        if db.query(models.OrdenCompra).filter(models.OrdenCompra.folio == oc["folio"]).first():
            skipped += 1
            continue
        prov = _proveedor_por_email(db, oc["proveedor_email"])
        if not prov:
            log.warning("Proveedor %s no encontrado, salto %s", oc["proveedor_email"], oc["folio"])
            continue

        subtotal = Decimal("0.00")
        detalles = []
        for sku, nombre, qty, costo in oc["lineas"]:
            costo_d = Decimal(costo)
            prod = _ensure_producto(db, sku, nombre, costo_d, oc["moneda"])
            subtotal += (costo_d * qty).quantize(Decimal("0.01"))
            detalles.append(models.DetalleCompra(
                producto_id=prod.id,
                cantidad=qty,
                costo_unitario=costo_d,
            ))

        iva = (subtotal * oc.get("iva_porcentaje", Decimal("0"))).quantize(Decimal("0.01"))
        total = (subtotal + iva).quantize(Decimal("0.01"))

        orden = models.OrdenCompra(
            folio=oc["folio"],
            proveedor_id=prov.id,
            fecha=oc["fecha"],
            total=total,
            estatus="confirmada",
            moneda=oc["moneda"],
            tipo_cambio=Decimal("1.0"),
            detalles=detalles,
        )
        db.add(orden)
        created += 1
    if not dry_run:
        db.commit()
    else:
        db.rollback()
    return created, skipped


# ---------- main ----------

def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--dry-run", action="store_true", help="No persiste cambios")
    args = ap.parse_args()

    if not LISTA_MATERIAL_FILE.exists():
        log.error("No existe %s", LISTA_MATERIAL_FILE)
        return 2
    if not REF_VTA_FILE.exists():
        log.warning("No existe %s — se omite taxonomía de marcas", REF_VTA_FILE)

    log.info("== Productos (LISTA_MATERIAL.xlsx) ==")
    productos = parse_lista_material(LISTA_MATERIAL_FILE)
    log.info("Filas leídas: %d", len(productos))

    log.info("== Taxonomía de marcas (REF.VTA) ==")
    marcas = parse_ref_vta(REF_VTA_FILE) if REF_VTA_FILE.exists() else []
    log.info("Abreviaturas leídas: %d", len(marcas))

    db = SessionLocal()
    try:
        created_p, updated_p = upsert_productos(db, productos, args.dry_run)
        log.info("Productos: %d creados, %d actualizados", created_p, updated_p)

        created_c, skipped_c = upsert_clientes(db, args.dry_run)
        log.info("Clientes: %d creados, %d ya existentes", created_c, skipped_c)

        created_pr, skipped_pr = upsert_proveedores(db, args.dry_run)
        log.info("Proveedores: %d creados, %d ya existentes", created_pr, skipped_pr)

        created_q, skipped_q = upsert_cotizaciones(db, args.dry_run)
        log.info("Cotizaciones de muestra: %d creadas, %d ya existentes", created_q, skipped_q)

        created_oc, skipped_oc = upsert_ordenes_compra(db, args.dry_run)
        log.info("Órdenes de compra de muestra: %d creadas, %d ya existentes", created_oc, skipped_oc)

        out = write_marca_taxonomy(marcas, args.dry_run)
        log.info("Taxonomía de marcas escrita en %s", out.relative_to(ROOT))

        if args.dry_run:
            log.info("--dry-run: nada fue persistido en la base de datos.")
        else:
            log.info("Listo.")
    finally:
        db.close()
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
